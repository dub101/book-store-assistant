import json
from pathlib import Path

import openpyxl

from book_store_assistant.export.workbook import apply_sheet_basics
from book_store_assistant.sources.models import SourceBookRecord
from book_store_assistant.sources.results import FetchResult

INTERMEDIATE_HEADERS = [
    "ISBN",
    "HasRecord",
    "Errors",
    "IssueCodes",
    "SourceName",
    "SourceURL",
    "Title",
    "Subtitle",
    "Author",
    "Editorial",
    "Synopsis",
    "Subject",
    "Categories",
    "CoverURL",
    "Language",
    "FieldSources",
]
INTERMEDIATE_SHEET_NAME = "FetchResults"


def _dump_json(value: object) -> str:
    return json.dumps(value, ensure_ascii=True)


def _load_json_list(value: object) -> list[str]:
    if not isinstance(value, str) or not value.strip():
        return []

    loaded = json.loads(value)
    if not isinstance(loaded, list):
        return []

    return [str(item) for item in loaded]


def _load_json_dict(value: object) -> dict[str, str]:
    if not isinstance(value, str) or not value.strip():
        return {}

    loaded = json.loads(value)
    if not isinstance(loaded, dict):
        return {}

    return {str(key): str(item) for key, item in loaded.items()}


def export_fetch_results(results: list[FetchResult], output_path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = INTERMEDIATE_SHEET_NAME
    sheet.append(INTERMEDIATE_HEADERS)

    for result in results:
        record = result.record
        sheet.append(
            [
                result.isbn,
                "yes" if record is not None else "no",
                _dump_json(result.errors),
                _dump_json(result.issue_codes),
                record.source_name if record is not None else None,
                str(record.source_url) if record is not None and record.source_url else None,
                record.title if record is not None else None,
                record.subtitle if record is not None else None,
                record.author if record is not None else None,
                record.editorial if record is not None else None,
                record.synopsis if record is not None else None,
                record.subject if record is not None else None,
                _dump_json(record.categories if record is not None else []),
                str(record.cover_url) if record is not None and record.cover_url else None,
                record.language if record is not None else None,
                _dump_json(record.field_sources if record is not None else {}),
            ]
        )

    apply_sheet_basics(sheet, freeze_panes="A2", wrap_columns=(10, 12))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def read_fetch_results(path: Path) -> list[FetchResult]:
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    results: list[FetchResult] = []

    for row in rows:
        (
            isbn,
            has_record,
            errors,
            issue_codes,
            source_name,
            source_url,
            title,
            subtitle,
            author,
            editorial,
            synopsis,
            subject,
            categories,
            cover_url,
            language,
            field_sources,
        ) = row

        record = None
        if has_record == "yes" and isinstance(isbn, str) and isinstance(source_name, str):
            record = SourceBookRecord(
                source_name=source_name,
                isbn=isbn,
                source_url=source_url,
                title=title,
                subtitle=subtitle,
                author=author,
                editorial=editorial,
                synopsis=synopsis,
                subject=subject,
                categories=_load_json_list(categories),
                cover_url=cover_url,
                language=language,
                field_sources=_load_json_dict(field_sources),
            )

        results.append(
            FetchResult(
                isbn=str(isbn),
                record=record,
                errors=_load_json_list(errors),
                issue_codes=_load_json_list(issue_codes),
            )
        )

    return results
