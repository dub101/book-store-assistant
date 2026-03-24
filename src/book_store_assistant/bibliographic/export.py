from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment

from book_store_assistant.bibliographic.models import BibliographicRecord
from book_store_assistant.resolution.results import ResolutionResult

UPLOAD_HEADERS = [
    "ISBN",
    "Title",
    "Subtitle",
    "Author",
    "Editorial",
    "Publisher",
]
REVIEW_HEADERS = [
    "ISBN",
    "Title",
    "Subtitle",
    "Author",
    "Editorial",
    "Publisher",
    "Status",
    "ReasonCode",
    "ValidatorConfidence",
    "ReviewNote",
]


def _validate_upload_row(row: list[str | None]) -> None:
    required_indexes = [0, 1, 3, 4, 5]
    for index in required_indexes:
        value = row[index]
        if value is None or not value.strip():
            raise ValueError(f"Upload row field '{UPLOAD_HEADERS[index]}' is required.")


def _upload_row(record: BibliographicRecord) -> list[str | None]:
    return [
        record.isbn,
        record.title,
        record.subtitle,
        record.author,
        record.editorial,
        record.publisher,
    ]


def _review_value(
    result: ResolutionResult,
    field_name: str,
) -> str | None:
    if result.candidate_record is not None and hasattr(result.candidate_record, field_name):
        value = getattr(result.candidate_record, field_name)
        if isinstance(value, str) and value.strip():
            return value

    if result.source_record is not None:
        value = getattr(result.source_record, field_name, None)
        if isinstance(value, str) and value.strip():
            return value

    if field_name == "publisher" and result.publisher_identity is not None:
        return result.publisher_identity.publisher_name

    if field_name == "editorial" and result.publisher_identity is not None:
        return result.publisher_identity.imprint_name or result.publisher_identity.publisher_name

    return None


def _review_row(result: ResolutionResult) -> list[str | None]:
    review_note = "; ".join(result.review_details) if result.review_details else None
    if result.validation_assessment is not None and result.validation_assessment.explanation:
        review_note = result.validation_assessment.explanation

    validator_confidence = (
        f"{result.validation_assessment.confidence:.2f}"
        if result.validation_assessment is not None
        else None
    )
    return [
        _review_value(result, "isbn"),
        _review_value(result, "title"),
        _review_value(result, "subtitle"),
        _review_value(result, "author"),
        _review_value(result, "editorial"),
        _review_value(result, "publisher"),
        "review",
        ", ".join(result.reason_codes) if result.reason_codes else None,
        validator_confidence,
        review_note,
    ]


def _apply_sheet_basics(
    sheet,
    *,
    freeze_panes: str,
    wrap_columns: tuple[int, int] | None = None,
) -> None:
    sheet.freeze_panes = freeze_panes
    sheet.auto_filter.ref = sheet.dimensions

    if wrap_columns is None:
        return

    min_column, max_column = wrap_columns
    for row in sheet.iter_rows(min_row=2, min_col=min_column, max_col=max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def export_upload_records(results: list[ResolutionResult], output_path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Upload"
    sheet.append(UPLOAD_HEADERS)

    for result in results:
        if result.record is None:
            continue
        if not isinstance(result.record, BibliographicRecord):
            raise TypeError("Stage 1 upload export requires bibliographic records.")
        row = _upload_row(result.record)
        _validate_upload_row(row)
        sheet.append(row)

    _apply_sheet_basics(sheet, freeze_panes="A2")
    workbook.save(output_path)


def export_review_rows(results: list[ResolutionResult], output_path: Path) -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Review"
    sheet.append(REVIEW_HEADERS)

    for result in results:
        if result.record is not None:
            continue
        sheet.append(_review_row(result))

    _apply_sheet_basics(sheet, freeze_panes="A2", wrap_columns=(10, 10))
    workbook.save(output_path)


def export_handoff_results(results: list[ResolutionResult], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    lines = [result.model_dump_json() for result in results]
    output_path.write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8")
