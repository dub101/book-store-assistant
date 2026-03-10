from pathlib import Path

import openpyxl

from book_store_assistant.enrichment.models import EnrichmentResult, GeneratedSynopsis
from book_store_assistant.export.schema import REVIEW_HEADERS, REVIEW_SHEET_NAME
from book_store_assistant.models import BookRecord
from book_store_assistant.resolution.results import ResolutionResult
from book_store_assistant.sources.models import SourceBookRecord


def test_review_rows_can_be_written_to_excel(tmp_path: Path) -> None:
    from book_store_assistant.export.review import export_review_rows

    output_file = tmp_path / "review.xlsx"
    source_record = SourceBookRecord(
        source_name="google_books + open_library",
        isbn="9780306406157",
        title="Example Title",
        subtitle="Example Subtitle",
        author="Example Author",
        editorial="Example Editorial",
        synopsis="Book description.",
        subject="FICCION",
        language="en",
        categories=["Fiction", "Literature"],
        cover_url="https://example.com/cover.jpg",
        field_sources={
            "title": "google_books",
            "editorial": "open_library",
            "categories": "google_books + open_library",
        },
    )
    results = [
        ResolutionResult(
            record=None,
            source_record=source_record,
            enrichment_result=EnrichmentResult(
                isbn="9780306406157",
                source_name="google_books",
                skipped_reason="generated_synopsis_rejected",
                evidence=[
                    {
                        "source_name": "google_books",
                        "evidence_type": "source_synopsis",
                        "evidence_origin": "direct_source_record",
                        "text": "Book description.",
                        "extraction_method": "source_synopsis_field",
                    }
                ],
                generated_synopsis=GeneratedSynopsis(
                    text="Resumen corto",
                    evidence_indexes=[0],
                    validation_flags=["generated_synopsis_too_short"],
                    raw_output_text=(
                        '{"text":"Resumen corto","language":"es","evidence_indexes":[0]}'
                    ),
                ),
            ),
            errors=[
                "Synopsis is missing.",
                "Synopsis came from google_books with language 'en'.",
            ],
            reason_codes=["MISSING_SYNOPSIS"],
            review_details=["Synopsis came from google_books with language 'en'."],
        ),
        ResolutionResult(
            record=BookRecord(
                isbn="0306406152",
                title="Resolved Title",
                author="Example Author",
                editorial="Example Editorial",
                synopsis="Resumen del libro.",
                subject="Narrativa",
            ),
            source_record=source_record,
            errors=[],
            reason_codes=[],
            review_details=[],
        ),
    ]

    export_review_rows(results, output_file)

    workbook = openpyxl.load_workbook(output_file)
    sheet = workbook.active

    assert sheet.title == REVIEW_SHEET_NAME
    assert [cell.value for cell in sheet[1]] == REVIEW_HEADERS
    assert sheet.cell(row=2, column=1).value == "9780306406157"
    assert sheet.cell(row=2, column=8).value == "FICCION"
    assert sheet.cell(row=2, column=9).value == "13"
    assert sheet.cell(row=2, column=10).value == "L0"
    assert sheet.cell(row=2, column=13).value == "Book description."
    assert "categories=google_books + open_library" in sheet.cell(row=2, column=14).value
    assert sheet.cell(row=2, column=15).value is None
    assert sheet.cell(row=2, column=16).value == "generated_synopsis_rejected"
    assert sheet.cell(row=2, column=17).value == "1"
    assert sheet.cell(row=2, column=18).value == "direct_source_record=1"
    assert sheet.cell(row=2, column=19).value == "generated_synopsis_too_short"
    assert sheet.cell(row=2, column=20).value == "Resumen corto"
    assert sheet.cell(row=2, column=21).value == (
        '{"text":"Resumen corto","language":"es","evidence_indexes":[0]}'
    )
    assert sheet.cell(row=2, column=22).value == "MISSING_SYNOPSIS"
    assert (
        "Synopsis came from google_books with language 'en'."
        in sheet.cell(row=2, column=23).value
    )
    assert sheet.cell(row=2, column=13).alignment.wrap_text is True
    assert sheet.cell(row=2, column=14).alignment.wrap_text is True
    assert sheet.cell(row=2, column=18).alignment.wrap_text is True
    assert sheet.cell(row=2, column=19).alignment.wrap_text is True
    assert sheet.cell(row=2, column=20).alignment.wrap_text is True
    assert sheet.cell(row=2, column=21).alignment.wrap_text is True
    assert sheet.cell(row=2, column=22).alignment.wrap_text is True
    assert sheet.cell(row=2, column=23).alignment.wrap_text is True
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:W2"
    assert sheet.max_row == 2
