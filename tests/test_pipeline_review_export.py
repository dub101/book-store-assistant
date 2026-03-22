from pathlib import Path

import openpyxl

from book_store_assistant.bibliographic.models import BibliographicRecord
from book_store_assistant.pipeline.review_export import export_unresolved_results
from book_store_assistant.resolution.models import RecordValidationAssessment
from book_store_assistant.resolution.results import ResolutionResult
from book_store_assistant.sources.models import SourceBookRecord


def test_export_unresolved_results_writes_only_unresolved_rows(tmp_path: Path) -> None:
    output_file = tmp_path / "review.xlsx"
    source_record = SourceBookRecord(source_name="google_books", isbn="9780306406157")
    results = [
        ResolutionResult(
            record=None,
            candidate_record=BibliographicRecord(
                isbn="9780306406157",
                title="Questionable Title",
                author="Questionable Author",
                editorial="Questionable Editorial",
                publisher="Questionable Publisher",
            ),
            source_record=source_record,
            validation_assessment=RecordValidationAssessment(
                accepted=False,
                confidence=0.42,
                explanation="Needs a quick human check.",
            ),
            errors=["LLM validation rejected the bibliographic record."],
            reason_codes=["VALIDATION_REJECTED"],
            review_details=["Needs a quick human check."],
        ),
        ResolutionResult(
            record=BibliographicRecord(
                isbn="0306406152",
                title="Example Title",
                author="Example Author",
                editorial="Example Editorial",
                publisher="Example Publisher",
            ),
            source_record=source_record,
            errors=[],
        ),
    ]

    export_unresolved_results(results, output_file)

    workbook = openpyxl.load_workbook(output_file)
    sheet = workbook.active

    assert sheet.max_row == 2
    assert sheet.cell(row=2, column=1).value == "9780306406157"
    assert sheet.cell(row=2, column=8).value == "VALIDATION_REJECTED"
    assert sheet.cell(row=2, column=9).value == "0.42"
    assert sheet.cell(row=2, column=10).value == "Needs a quick human check."


def test_export_unresolved_results_preserves_isbn_for_fetch_failures(tmp_path: Path) -> None:
    output_file = tmp_path / "review.xlsx"
    results = [
        ResolutionResult(
            record=None,
            source_record=SourceBookRecord(
                source_name="fetch_error",
                isbn="9780306406157",
            ),
            errors=["google_books: No Google Books match found."],
        )
    ]

    export_unresolved_results(results, output_file)

    workbook = openpyxl.load_workbook(output_file)
    sheet = workbook.active

    assert sheet.max_row == 2
    assert sheet.cell(row=2, column=1).value == "9780306406157"
    assert sheet.cell(row=2, column=7).value == "review"
    assert sheet.cell(row=2, column=10).value is None
