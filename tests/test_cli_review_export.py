from pathlib import Path
from unittest.mock import patch

import openpyxl
from typer.testing import CliRunner

from book_store_assistant.bibliographic.models import BibliographicRecord
from book_store_assistant.cli import app
from book_store_assistant.pipeline.contracts import ISBNInput
from book_store_assistant.pipeline.process_results import ProcessResult
from book_store_assistant.pipeline.results import InputReadResult
from book_store_assistant.resolution.models import RecordValidationAssessment
from book_store_assistant.resolution.results import ResolutionResult
from book_store_assistant.sources.models import SourceBookRecord

runner = CliRunner()

EXPECTED_REVIEW_HEADERS = [
    "ISBN", "Title", "Subtitle", "Author", "Editorial",
    "Status", "ReasonCode", "ValidatorConfidence", "ReviewNote",
]


@patch("book_store_assistant.cli.process_isbn_file")
def test_cli_can_export_review_rows(mock_process_isbn_file, tmp_path: Path) -> None:
    input_file = tmp_path / "isbns.csv"
    review_file = tmp_path / "review.xlsx"
    input_file.write_text("9780306406157\n", encoding="utf-8")

    mock_process_isbn_file.return_value = ProcessResult(
        input_result=InputReadResult(
            valid_inputs=[ISBNInput(isbn="9780306406157")],
            invalid_values=[],
        ),
        fetch_results=[],
        resolution_results=[
            ResolutionResult(
                record=None,
                candidate_record=BibliographicRecord(
                    isbn="9780306406157",
                    title="Example Title",
                    author="Example Author",
                    editorial="Example Editorial",
                ),
                source_record=SourceBookRecord(
                    source_name="google_books",
                    isbn="9780306406157",
                    title="Example Title",
                    author="Example Author",
                    editorial="Example Editorial",
                ),
                validation_assessment=RecordValidationAssessment(
                    accepted=False,
                    confidence=0.42,
                    explanation="Author needs a manual check.",
                ),
                errors=["LLM validation rejected the bibliographic record."],
                reason_codes=["VALIDATION_REJECTED"],
                review_details=["Author needs a manual check."],
            ),
            ResolutionResult(
                record=BibliographicRecord(
                    isbn="0306406152",
                    title="Resolved Title",
                    subtitle="Resolved Subtitle",
                    author="Example Author",
                    editorial="Example Editorial",
                    synopsis="Sinopsis resuelta.",
                    subject="NOVELA",
                    subject_code="20",
                ),
                source_record=None,
                errors=[],
            ),
        ],
    )

    result = runner.invoke(app, [str(input_file), "--review-output", str(review_file)])

    assert result.exit_code == 0
    workbook = openpyxl.load_workbook(review_file)
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]] == EXPECTED_REVIEW_HEADERS
    row2 = [sheet.cell(row=2, column=i).value for i in range(1, 10)]
    assert row2[0] == "9780306406157"
    assert row2[1] == "Example Title"
    assert row2[4] == "Example Editorial"
    assert row2[5] == "review"
    assert row2[6] == "VALIDATION_REJECTED"
    assert row2[7] == "0.42"
    assert row2[8] == "Author needs a manual check."


@patch("book_store_assistant.cli.process_isbn_file")
def test_cli_review_export_uses_exact_requested_path(
    mock_process_isbn_file,
    tmp_path: Path,
) -> None:
    input_file = tmp_path / "isbns.csv"
    review_file = tmp_path / "review.xlsx"
    input_file.write_text("9780306406157\n", encoding="utf-8")

    mock_process_isbn_file.return_value = ProcessResult(
        input_result=InputReadResult(
            valid_inputs=[ISBNInput(isbn="9780306406157")],
            invalid_values=[],
        ),
        fetch_results=[],
        resolution_results=[
            ResolutionResult(
                record=None,
                source_record=SourceBookRecord(
                    source_name="google_books",
                    isbn="9780306406157",
                    title="Example Title",
                ),
                errors=["Title is missing."],
                reason_codes=["MISSING_TITLE"],
            ),
        ],
    )

    result = runner.invoke(app, [str(input_file), "--review-output", str(review_file)])

    assert result.exit_code == 0
    assert review_file.exists()
    assert openpyxl.load_workbook(review_file).active.title == "Review"


@patch("book_store_assistant.cli.process_isbn_file")
def test_cli_review_export_does_not_create_implicit_output_variants(
    mock_process_isbn_file,
    tmp_path: Path,
) -> None:
    input_file = tmp_path / "isbns.csv"
    review_file = tmp_path / "review.xlsx"
    input_file.write_text("9780306406157\n", encoding="utf-8")

    mock_process_isbn_file.return_value = ProcessResult(
        input_result=InputReadResult(
            valid_inputs=[ISBNInput(isbn="9780306406157")],
            invalid_values=[],
        ),
        fetch_results=[],
        resolution_results=[
            ResolutionResult(
                record=None,
                source_record=SourceBookRecord(
                    source_name="google_books",
                    isbn="9780306406157",
                    title="Example Title",
                ),
                errors=["Title is missing."],
                reason_codes=["MISSING_TITLE"],
            ),
        ],
    )

    result = runner.invoke(app, [str(input_file), "--review-output", str(review_file)])

    assert result.exit_code == 0
    assert review_file.exists()
    assert not (tmp_path / "review.rules-only.xlsx").exists()
