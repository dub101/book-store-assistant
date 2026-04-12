import json
from pathlib import Path

import openpyxl

from book_store_assistant.bibliographic.export import (
    REVIEW_HEADERS,
    UPLOAD_HEADERS,
    export_handoff_results,
    export_review_rows,
    export_upload_records,
)
from book_store_assistant.bibliographic.models import BibliographicRecord
from book_store_assistant.resolution.models import RecordValidationAssessment
from book_store_assistant.resolution.results import ResolutionResult
from book_store_assistant.sources.models import SourceBookRecord


def test_export_upload_records_writes_expected_columns_and_row(tmp_path: Path) -> None:
    output_file = tmp_path / "upload.xlsx"
    results = [
        ResolutionResult(
            record=BibliographicRecord(
                isbn="9780306406157",
                title="Example Title",
                subtitle="Example Subtitle",
                author="Example Author",
                editorial="Example Editorial",
                synopsis="Sinopsis de ejemplo en español.",
                subject="NOVELA",
                subject_code="20",
            ),
            candidate_record=BibliographicRecord(
                isbn="9780306406157",
                title="Example Title",
                subtitle="Example Subtitle",
                author="Example Author",
                editorial="Example Editorial",
                synopsis="Sinopsis de ejemplo en español.",
                subject="NOVELA",
                subject_code="20",
            ),
            source_record=SourceBookRecord(source_name="google_books", isbn="9780306406157"),
            errors=[],
            reason_codes=[],
            review_details=[],
        )
    ]

    export_upload_records(results, output_file)

    workbook = openpyxl.load_workbook(output_file)
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]] == UPLOAD_HEADERS
    assert sheet.cell(row=2, column=1).value == "9780306406157"
    assert sheet.cell(row=2, column=2).value == "Example Title"
    assert sheet.cell(row=2, column=4).value == "Example Author"
    assert sheet.cell(row=2, column=5).value == "Example Editorial"


def test_export_review_rows_and_handoff_write_expected_outputs(tmp_path: Path) -> None:
    review_file = tmp_path / "review.xlsx"
    handoff_file = tmp_path / "handoff.jsonl"
    result = ResolutionResult(
        record=None,
        candidate_record=BibliographicRecord(
            isbn="9780306406157",
            title="Questionable Title",
            author="Questionable Author",
            editorial="Questionable Editorial",
        ),
        source_record=SourceBookRecord(
            source_name="google_books",
            isbn="9780306406157",
            title="Questionable Title",
            author="Questionable Author",
            editorial="Questionable Editorial",
        ),
        validation_assessment=RecordValidationAssessment(
            accepted=False,
            confidence=0.45,
            explanation="Title is not well supported by the source evidence.",
        ),
        errors=["LLM validation rejected the bibliographic record."],
        reason_codes=["VALIDATION_REJECTED"],
        review_details=["Title is not well supported by the source evidence."],
        path_summary={"first_material_gain_stage": "llm_enrichment"},
    )

    export_review_rows([result], review_file)
    export_handoff_results([result], handoff_file)

    workbook = openpyxl.load_workbook(review_file)
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]] == REVIEW_HEADERS
    assert sheet.cell(row=2, column=1).value == "9780306406157"
    assert sheet.cell(row=2, column=7).value == "VALIDATION_REJECTED"
    assert sheet.cell(row=2, column=8).value == "0.45"
    payload = json.loads(handoff_file.read_text(encoding="utf-8").strip())
    assert payload["path_summary"]["first_material_gain_stage"] == "llm_enrichment"
