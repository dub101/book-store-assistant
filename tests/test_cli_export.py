from pathlib import Path
from unittest.mock import patch

import openpyxl
from typer.testing import CliRunner

from book_store_assistant.bibliographic.models import BibliographicRecord
from book_store_assistant.cli import app
from book_store_assistant.pipeline.contracts import ISBNInput
from book_store_assistant.pipeline.process_results import ProcessResult
from book_store_assistant.pipeline.results import InputReadResult
from book_store_assistant.resolution.results import ResolutionResult

runner = CliRunner()

EXPECTED_UPLOAD_HEADERS = [
    "ISBN", "Title", "Subtitle", "Author", "Editorial",
]


@patch("book_store_assistant.cli.process_isbn_file")
def test_cli_export_writes_resolved_records(mock_process_isbn_file, tmp_path: Path) -> None:
    input_file = tmp_path / "isbns.csv"
    output_file = tmp_path / "books.xlsx"
    input_file.write_text("9780306406157\n", encoding="utf-8")

    mock_process_isbn_file.return_value = ProcessResult(
        input_result=InputReadResult(
            valid_inputs=[ISBNInput(isbn="9780306406157")],
            invalid_values=[],
        ),
        fetch_results=[],
        resolution_results=[
            ResolutionResult(
                record=BibliographicRecord(
                    isbn="9780306406157",
                    title="Example Title",
                    subtitle="Example Subtitle",
                    author="Example Author",
                    editorial="Example Editorial",
                    synopsis="Sinopsis de ejemplo.",
                    subject="NOVELA",
                    subject_code="20",
                ),
                source_record=None,
                errors=[],
            )
        ],
    )

    result = runner.invoke(app, [str(input_file), "--output", str(output_file)])

    assert result.exit_code == 0
    workbook = openpyxl.load_workbook(output_file)
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]] == EXPECTED_UPLOAD_HEADERS
    assert sheet.cell(row=2, column=1).value == "9780306406157"
    assert sheet.cell(row=2, column=2).value == "Example Title"
    assert sheet.cell(row=2, column=4).value == "Example Author"
    assert sheet.cell(row=2, column=5).value == "Example Editorial"


@patch("book_store_assistant.cli.process_isbn_file")
def test_cli_export_writes_default_handoff_next_to_upload_output(
    mock_process_isbn_file,
    tmp_path: Path,
) -> None:
    input_file = tmp_path / "isbns.csv"
    output_file = tmp_path / "books.xlsx"
    expected_handoff_file = tmp_path / "books.handoff.jsonl"
    input_file.write_text("9780306406157\n", encoding="utf-8")

    mock_process_isbn_file.return_value = ProcessResult(
        input_result=InputReadResult(
            valid_inputs=[ISBNInput(isbn="9780306406157")],
            invalid_values=[],
        ),
        fetch_results=[],
        resolution_results=[
            ResolutionResult(
                record=BibliographicRecord(
                    isbn="9780306406157",
                    title="Example Title",
                    author="Example Author",
                    editorial="Example Editorial",
                    synopsis="Sinopsis de ejemplo.",
                    subject="NOVELA",
                    subject_code="20",
                ),
                source_record=None,
                errors=[],
            )
        ],
    )

    result = runner.invoke(app, [str(input_file), "--output", str(output_file)])

    assert result.exit_code == 0
    assert output_file.exists()
    assert expected_handoff_file.exists()
    assert '"isbn":"9780306406157"' in expected_handoff_file.read_text(encoding="utf-8")


@patch("book_store_assistant.cli.process_isbn_file")
def test_cli_export_respects_explicit_handoff_output_path(
    mock_process_isbn_file,
    tmp_path: Path,
) -> None:
    input_file = tmp_path / "isbns.csv"
    output_file = tmp_path / "books.xlsx"
    handoff_file = tmp_path / "custom-name.jsonl"
    input_file.write_text("9780306406157\n", encoding="utf-8")

    mock_process_isbn_file.return_value = ProcessResult(
        input_result=InputReadResult(
            valid_inputs=[ISBNInput(isbn="9780306406157")],
            invalid_values=[],
        ),
        fetch_results=[],
        resolution_results=[
            ResolutionResult(
                record=BibliographicRecord(
                    isbn="9780306406157",
                    title="Example Title",
                    author="Example Author",
                    editorial="Example Editorial",
                    synopsis="Sinopsis de ejemplo.",
                    subject="NOVELA",
                    subject_code="20",
                ),
                source_record=None,
                errors=[],
            )
        ],
    )

    result = runner.invoke(
        app,
        [str(input_file), "--output", str(output_file), "--handoff-output", str(handoff_file)],
    )

    assert result.exit_code == 0
    assert output_file.exists()
    assert handoff_file.exists()
    assert not (tmp_path / "books.handoff.jsonl").exists()
