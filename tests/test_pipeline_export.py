from pathlib import Path

import openpyxl

from book_store_assistant.models import BookRecord
from book_store_assistant.pipeline.export import export_resolved_records
from book_store_assistant.resolution.results import ResolutionResult
from book_store_assistant.sources.models import SourceBookRecord


def test_export_resolved_records_writes_only_resolved_rows(tmp_path: Path) -> None:
    output_file = tmp_path / "books.xlsx"
    source_record = SourceBookRecord(source_name="google_books", isbn="9780306406157")
    results = [
        ResolutionResult(
            record=None,
            source_record=source_record,
            errors=["Synopsis is missing."],
        ),
        ResolutionResult(
            record=BookRecord(
                isbn="0306406152",
                title="Example Title",
                author="Example Author",
                editorial="Example Editorial",
                synopsis="Resumen del libro.",
                subject="FICCION",
            ),
            source_record=source_record,
            errors=[],
        ),
    ]

    export_resolved_records(results, output_file)

    workbook = openpyxl.load_workbook(output_file)
    sheet = workbook.active

    assert sheet.max_row == 2
    assert sheet.cell(row=2, column=1).value == "0306406152"
