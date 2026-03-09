from pathlib import Path

import openpyxl

from book_store_assistant.export.excel import export_books
from book_store_assistant.export.schema import BOOKS_HEADERS, BOOKS_SHEET_NAME
from book_store_assistant.models import BookRecord


def test_export_books_writes_expected_columns_and_row(tmp_path: Path) -> None:
    output_file = tmp_path / "books.xlsx"
    records = [
        BookRecord(
            isbn="9780306406157",
            title="Example Title",
            subtitle="Example Subtitle",
            author="Example Author",
            editorial="Example Editorial",
            synopsis="Resumen del libro.",
            subject="Narrativa",
            cover_url="https://example.com/cover.jpg",
        )
    ]

    export_books(records, output_file)

    workbook = openpyxl.load_workbook(output_file)
    sheet = workbook.active

    assert sheet.title == BOOKS_SHEET_NAME
    assert [cell.value for cell in sheet[1]] == BOOKS_HEADERS
    assert sheet.cell(row=2, column=1).value == "9780306406157"
    assert sheet.cell(row=2, column=6).value == "Resumen del libro."
    assert sheet.cell(row=2, column=6).alignment.wrap_text is True
    assert sheet.cell(row=2, column=6).alignment.vertical == "top"
    assert sheet.cell(row=2, column=8).value == "https://example.com/cover.jpg"
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:H2"
