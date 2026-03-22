from pathlib import Path
from unittest.mock import patch

import openpyxl

from book_store_assistant.bibliographic.export import (
    export_review_rows,
    export_upload_records,
)
from book_store_assistant.config import AppConfig
from book_store_assistant.pipeline.input import read_isbn_inputs
from book_store_assistant.pipeline.service import process_isbn_file
from book_store_assistant.resolution.models import RecordValidationAssessment
from book_store_assistant.sources.models import SourceBookRecord
from book_store_assistant.sources.results import FetchResult

PROJECT_ROOT = Path(__file__).resolve().parents[1]
SAMPLE_1_PATH = PROJECT_ROOT / "data" / "input" / "sample_1.csv"
SAMPLE_2_PATH = PROJECT_ROOT / "data" / "input" / "sample_2.csv"


class AcceptingValidator:
    def validate(self, source_record, candidate_record, publisher_identity=None):
        return RecordValidationAssessment(accepted=True, confidence=0.97)


class FixtureBatchSource:
    def __init__(self, sample_name: str, fixture_isbns: list[str]) -> None:
        self.fixture_results = self._build_fixture_results(sample_name, fixture_isbns)

    def fetch(self, isbn: str) -> FetchResult:
        return self.fixture_results[isbn]

    def _build_fixture_results(
        self,
        sample_name: str,
        fixture_isbns: list[str],
    ) -> dict[str, FetchResult]:
        if sample_name == "sample_1":
            return _build_sample_1_fixture_results(fixture_isbns)

        return _build_sample_2_fixture_results(fixture_isbns)


def _build_fixture_record(
    isbn: str,
    *,
    author: str | None = "Fixture Author",
    editorial: str | None = "Fixture Editorial",
) -> SourceBookRecord:
    field_sources = {
        "title": "fixture_source",
        "author": "fixture_source",
        "editorial": "fixture_source",
    }

    return SourceBookRecord(
        source_name="fixture_source",
        isbn=isbn,
        title=f"Fixture Title {isbn[-4:]}",
        subtitle=f"Fixture Subtitle {isbn[-2:]}",
        author=author,
        editorial=editorial,
        field_sources=field_sources,
    )


def _build_fetch_error(isbn: str) -> FetchResult:
    return FetchResult(
        isbn=isbn,
        record=None,
        errors=["fixture_source: simulated fetch failure"],
        issue_codes=["FIXTURE_SOURCE:FIXTURE_FETCH_FAILURE"],
    )


def _build_sample_1_fixture_results(fixture_isbns: list[str]) -> dict[str, FetchResult]:
    fetch_error_indexes = {8}
    results: dict[str, FetchResult] = {}

    for index, isbn in enumerate(fixture_isbns):
        if index in fetch_error_indexes:
            results[isbn] = _build_fetch_error(isbn)
            continue

        record = _build_fixture_record(isbn)
        results[isbn] = FetchResult(isbn=isbn, record=record, errors=[])

    return results


def _build_sample_2_fixture_results(fixture_isbns: list[str]) -> dict[str, FetchResult]:
    results: dict[str, FetchResult] = {}

    for index, isbn in enumerate(fixture_isbns):
        if index < 10:
            record = _build_fixture_record(isbn)
        elif index < 20:
            record = _build_fixture_record(isbn)
        elif index < 25:
            results[isbn] = _build_fetch_error(isbn)
            continue
        elif index < 30:
            record = _build_fixture_record(isbn)
        elif index < 35:
            record = _build_fixture_record(
                isbn,
                editorial=None,
            )
        elif index < 40:
            record = _build_fixture_record(
                isbn,
                author=None,
            )
        elif index < 45:
            record = _build_fixture_record(isbn)
        else:
            record = _build_fixture_record(isbn)

        results[isbn] = FetchResult(isbn=isbn, record=record, errors=[])

    return results


def test_sample_1_batch_regression_in_stage_1_mode(tmp_path: Path) -> None:
    fixture_isbns = [item.isbn for item in read_isbn_inputs(SAMPLE_1_PATH).valid_inputs]
    with patch(
        "book_store_assistant.pipeline.service.build_default_record_quality_validator",
        return_value=AcceptingValidator(),
    ):
        result = process_isbn_file(
            SAMPLE_1_PATH,
            source=FixtureBatchSource("sample_1", fixture_isbns),
            config=AppConfig(
                retailer_page_lookup_enabled=False,
                publisher_page_lookup_enabled=False,
                web_search_fallback_enabled=False,
                publisher_page_timeout_seconds=0.01,
                publisher_page_max_retries=0,
                publisher_page_backoff_seconds=0.0,
            ),
        )

    resolved_count = sum(1 for item in result.resolution_results if item.record is not None)
    unresolved_count = sum(1 for item in result.resolution_results if item.record is None)

    assert len(result.input_result.valid_inputs) == 10
    assert sum(1 for item in result.fetch_results if item.record is not None) == 9
    assert resolved_count == 9
    assert unresolved_count == 1

    resolved_output = tmp_path / "sample_1_books.xlsx"
    review_output = tmp_path / "sample_1_review.xlsx"
    export_upload_records(result.resolution_results, resolved_output)
    export_review_rows(result.resolution_results, review_output)

    resolved_sheet = openpyxl.load_workbook(resolved_output).active
    review_sheet = openpyxl.load_workbook(review_output).active

    assert resolved_sheet.max_row == 10
    assert review_sheet.max_row == 2


def test_sample_2_batch_regression_in_stage_1_mode(tmp_path: Path) -> None:
    fixture_isbns = [item.isbn for item in read_isbn_inputs(SAMPLE_2_PATH).valid_inputs]
    with patch(
        "book_store_assistant.pipeline.service.build_default_record_quality_validator",
        return_value=AcceptingValidator(),
    ):
        result = process_isbn_file(
            SAMPLE_2_PATH,
            source=FixtureBatchSource("sample_2", fixture_isbns),
            config=AppConfig(
                retailer_page_lookup_enabled=False,
                publisher_page_lookup_enabled=False,
                web_search_fallback_enabled=False,
                publisher_page_timeout_seconds=0.01,
                publisher_page_max_retries=0,
                publisher_page_backoff_seconds=0.0,
            ),
        )

    resolved_count = sum(1 for item in result.resolution_results if item.record is not None)
    unresolved_count = sum(1 for item in result.resolution_results if item.record is None)

    assert len(result.input_result.valid_inputs) == 49
    assert result.input_result.invalid_values == ["9788449333830"]
    assert sum(1 for item in result.fetch_results if item.record is not None) == 44
    assert resolved_count == 34
    assert unresolved_count == 15

    resolved_output = tmp_path / "sample_2_books.xlsx"
    review_output = tmp_path / "sample_2_review.xlsx"
    export_upload_records(result.resolution_results, resolved_output)
    export_review_rows(result.resolution_results, review_output)

    resolved_sheet = openpyxl.load_workbook(resolved_output).active
    review_sheet = openpyxl.load_workbook(review_output).active

    assert resolved_sheet.max_row == 35
    assert review_sheet.max_row == 16
