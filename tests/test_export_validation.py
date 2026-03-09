import openpyxl

from book_store_assistant.export.schema import (
    BOOKS_HEADERS,
    BOOKS_SHEET_NAME,
    REVIEW_HEADERS,
    REVIEW_SHEET_NAME,
)
from book_store_assistant.validation.export import (
    validate_books_sheet,
    validate_review_sheet,
)


def test_validate_books_sheet_accepts_expected_contract() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = BOOKS_SHEET_NAME
    sheet.append(BOOKS_HEADERS)
    sheet.freeze_panes = "A2"

    errors = validate_books_sheet(sheet)

    assert errors == []


def test_validate_books_sheet_reports_contract_mismatches() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["ISBN", "Title"])
    sheet.freeze_panes = "B2"

    errors = validate_books_sheet(sheet)

    assert f"Books sheet title must be '{BOOKS_SHEET_NAME}', got 'Sheet1'." in errors
    assert f"Books sheet headers must be {BOOKS_HEADERS}, got ['ISBN', 'Title']." in errors
    assert "Books sheet freeze panes must be 'A2', got 'B2'." in errors


def test_validate_review_sheet_accepts_expected_contract() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = REVIEW_SHEET_NAME
    sheet.append(REVIEW_HEADERS)
    sheet.freeze_panes = "A2"

    errors = validate_review_sheet(sheet)

    assert errors == []


def test_validate_review_sheet_reports_contract_mismatches() -> None:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(["ISBN", "Title"])
    sheet.freeze_panes = "B2"

    errors = validate_review_sheet(sheet)

    assert f"Review sheet title must be '{REVIEW_SHEET_NAME}', got 'Sheet1'." in errors
    assert f"Review sheet headers must be {REVIEW_HEADERS}, got ['ISBN', 'Title']." in errors
    assert "Review sheet freeze panes must be 'A2', got 'B2'." in errors
